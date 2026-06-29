"""
rank.py — Main ranking script. Completes in <5 minutes on CPU.

Run:
  python rank.py --input candidates.jsonl --precomputed precomputed/ --output submission.csv

Outputs both submission.csv (for validate_submission.py) and submission.xlsx (for portal upload).
"""

import argparse
import json
import os
import time
import csv
import numpy as np
from tqdm import tqdm
from features import structured_score, behavioral_score, final_score


def load_candidates(path: str):
    candidates = []
    ext = os.path.splitext(path)[1].lower()
    with open(path, "r", encoding="utf-8") as f:
        if ext == ".jsonl":
            for line in f:
                line = line.strip()
                if line:
                    candidates.append(json.loads(line))
        else:
            candidates = json.load(f)
    return candidates


def write_xlsx(rows, path):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ranked Candidates"
        ws.append(["candidate_id", "rank", "score", "reasoning"])
        for row in rows:
            ws.append(row)
        # Style header
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="1E2761", end_color="1E2761", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)
        wb.save(path)
        print(f"XLSX saved: {path}")
    except ImportError:
        print("openpyxl not installed — skipping XLSX. Run: pip install openpyxl")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="candidates.jsonl")
    parser.add_argument("--precomputed", default="precomputed")
    parser.add_argument("--output", default="submission.csv")
    parser.add_argument("--top-k", type=int, default=100)
    args = parser.parse_args()

    t0 = time.time()

    print("Loading precomputed embeddings...")
    embeddings = np.load(os.path.join(args.precomputed, "embeddings.npy"))
    candidate_ids = np.load(os.path.join(args.precomputed, "candidate_ids.npy"))
    jd_embedding = np.load(os.path.join(args.precomputed, "jd_embedding.npy"))
    print(f"  {len(embeddings):,} candidates | {time.time()-t0:.1f}s")

    print("Computing semantic scores...")
    semantic_scores = embeddings @ jd_embedding
    s_min, s_max = semantic_scores.min(), semantic_scores.max()
    semantic_scores = (semantic_scores - s_min) / (s_max - s_min + 1e-9)
    print(f"  Done | {time.time()-t0:.1f}s")

    print("Loading candidates...")
    candidates = load_candidates(args.input)
    print(f"  {len(candidates):,} candidates | {time.time()-t0:.1f}s")

    id_to_idx = {cid: i for i, cid in enumerate(candidate_ids)}

    print("Scoring all candidates...")
    results = []
    for c in tqdm(candidates):
        cid = c["candidate_id"]
        idx = id_to_idx.get(cid)
        if idx is None:
            continue
        sem = float(semantic_scores[idx])
        score, reasoning = final_score(c, semantic=sem)
        results.append((score, cid, reasoning))

    print(f"  Scoring done | {time.time()-t0:.1f}s")

    results.sort(key=lambda x: x[0], reverse=True)

    # Tie-break: same score → sort by candidate_id ascending (per validation rules)
    results.sort(key=lambda x: (-x[0], x[1]))

    top_k = results[:args.top_k]

    # Write CSV
    rows = []
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])
        for rank, (score, cid, reasoning) in enumerate(top_k, 1):
            row = [cid, rank, f"{score:.4f}", reasoning]
            writer.writerow(row)
            rows.append(row)
    print(f"CSV saved: {args.output}")

    # Write XLSX
    xlsx_path = args.output.replace(".csv", ".xlsx")
    write_xlsx(rows, xlsx_path)

    total = time.time() - t0
    print(f"\nCompleted in {total:.1f}s ({total/60:.1f} min)")
    print(f"\nTop 5 candidates:")
    for rank, (score, cid, reasoning) in enumerate(top_k[:5], 1):
        print(f"  {rank}. {cid} | {score:.4f} | {reasoning[:90]}")


if __name__ == "__main__":
    main()
